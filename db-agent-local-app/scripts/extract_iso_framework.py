from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract ISO 27001 workbook rows into JSON.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize(value) for value in rows[0]]

    records = []
    for index, row in enumerate(rows[1:], start=2):
      values = {headers[column]: normalize(cell) for column, cell in enumerate(row[: len(headers)])}
      identifier = values.get("Identifier", "")
      if not identifier:
        continue
      record = {
        "row_number": index,
        "identifier": identifier,
        "requirement_name": values.get("Requirement Name", ""),
        "requirement_control": values.get("Requirement Control", ""),
        "control_descriptive": values.get("Control (Descriptive)", ""),
        "evidence_task": values.get("Evidence Task", ""),
        "evidence_description": values.get("Evidence Description", ""),
      }
      record["search_text"] = " ".join(
        part for part in [
          record["identifier"],
          record["requirement_name"],
          record["requirement_control"],
          record["control_descriptive"],
          record["evidence_task"],
        ]
        if part
      ).lower()
      records.append(record)

    payload = {
      "source_workbook": str(input_path.resolve()),
      "sheet_name": sheet.title,
      "record_count": len(records),
      "requirements": records,
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"status": "success", "output": str(output_path.resolve()), "record_count": len(records)}))


if __name__ == "__main__":
    main()
