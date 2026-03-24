from __future__ import annotations

import argparse
import csv
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


SECTION_FILES = {
    "onboarding": ("Client Details", "onboarding.json"),
    "policy_generation": ("Policies and Procedures", "policies.json"),
    "risk_assessment": ("Risk Assessments", "risk-assessments.json"),
    "vendor_risk": ("Vendor Assessments", "vendor-assessments.json"),
    "control_mapping": ("Control Mapping", "controls.json"),
    "audit_qa": ("Audit QA", "audit-qa.json"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate DB Agent client exports.")
    parser.add_argument("--client-id", required=True)
    parser.add_argument("--client-root", required=True)
    parser.add_argument("--export-root", required=True)
    return parser.parse_args()


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat()


def read_json(path: Path, default):
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8-sig") as handle:
        raw = handle.read().strip()
    if not raw:
        return default
    return json.loads(raw)


def client_section_path(client_root: Path, key: str) -> Path:
    folder, filename = SECTION_FILES[key]
    return client_root / folder / filename


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(sanitize_text(item) for item in value if item not in (None, ""))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def clean_inline_text(value) -> str:
    text = sanitize_text(value).replace("\r", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").lower()
    return cleaned or "policy"


def split_sentences(text: str) -> list[str]:
    """Split a paragraph into bullet sentences at '. Capital' boundaries."""
    parts = re.split(r"\.\s+(?=[A-Z])", text)
    result = []
    for p in parts:
        p = p.strip().rstrip(".") + "."
        if len(p) > 16:
            result.append(p)
    return result if len(result) >= 2 else []


def write_csv(path: Path, rows: list[dict], headers: list[str]) -> str:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: sanitize_text(row.get(header, "")) for header in headers})
    return str(path.resolve())


def write_xlsx(path: Path, rows: list[dict], headers: list[str], sheet_name: str) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([sanitize_text(row.get(header, "")) for header in headers])
    for column_cells in sheet.columns:
        length = max(len(sanitize_text(cell.value)) for cell in column_cells if cell.value is not None)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 14), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)
    return str(path.resolve())


def build_policy_pdf(path: Path, client_name: str, policies: list[dict], generated_at: str, scope_text: str) -> str:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PolicyPackTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "PolicyPackSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#475569"),
        spaceAfter=4,
    )
    policy_title_style = ParagraphStyle(
        "PolicyTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=8,
    )
    section_heading_style = ParagraphStyle(
        "PolicySectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#0F172A"),
        spaceBefore=8,
        spaceAfter=6,
    )
    numbered_heading_style = ParagraphStyle(
        "PolicyNumberedHeading",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#0D7377"),
        spaceBefore=8,
        spaceAfter=2,
    )
    body_style = ParagraphStyle(
        "PolicyBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=4,
    )
    muted_style = ParagraphStyle(
        "PolicyMuted",
        parent=body_style,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#64748B"),
    )
    bullet_style = ParagraphStyle(
        "PolicyBullet",
        parent=body_style,
        leftIndent=14,
        firstLineIndent=-10,
        bulletIndent=0,
        spaceAfter=4,
    )

    def on_page(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.setFont("Helvetica", 8)
        canvas.line(document.leftMargin, letter[1] - 42, letter[0] - document.rightMargin, letter[1] - 42)
        canvas.drawString(document.leftMargin, letter[1] - 34, client_name)
        canvas.drawRightString(letter[0] - document.rightMargin, letter[1] - 34, "DB Agent Policy Pack")
        canvas.line(document.leftMargin, 36, letter[0] - document.rightMargin, 36)
        canvas.drawString(document.leftMargin, 24, f"Generated {generated_at}")
        canvas.drawRightString(letter[0] - document.rightMargin, 24, f"Page {document.page}")
        canvas.restoreState()

    def add_policy_text_block(story: list, text: str) -> None:
        for raw_line in sanitize_text(text).replace("\r", "").split("\n"):
            line = raw_line.strip()
            if not line:
                story.append(Spacer(1, 4))
                continue

            # Old inline format: "1.1 Heading. Content paragraph..."
            inline = re.match(
                r"^(\d+\.\d+(?:\.\d+)?\s+[A-Za-z][^.\r\n]{2,55})\.\s+([A-Z].{15,})$",
                line,
            )
            if inline:
                story.append(Paragraph(escape(inline.group(1)), numbered_heading_style))
                story.append(Spacer(1, 2))
                sentences = split_sentences(inline.group(2))
                if sentences:
                    for s in sentences:
                        story.append(Paragraph(f"• {escape(s)}", bullet_style))
                else:
                    story.append(Paragraph(escape(inline.group(2)), body_style))
                continue

            safe_line = escape(line)

            # Main section heading "1. Title" (not "1.1")
            if re.match(r"^\d+\.\s+\S", line) and not re.match(r"^\d+\.\d+", line):
                story.append(Paragraph(safe_line, section_heading_style))
            # Subsection heading already on its own line "1.1 Title"
            elif re.match(r"^\d+\.\d+(\.\d+)?\s+\S", line) and len(line) < 80:
                story.append(Paragraph(safe_line, numbered_heading_style))
                story.append(Spacer(1, 2))
            elif line.startswith("- "):
                story.append(Paragraph(f"• {escape(line[2:].strip())}", bullet_style))
            else:
                # Content paragraph — bullet if multiple sentences
                sentences = split_sentences(line)
                if sentences:
                    for s in sentences:
                        story.append(Paragraph(f"• {escape(s)}", bullet_style))
                else:
                    story.append(Paragraph(safe_line, body_style))

    story = [
        Paragraph("DB Agent Policy Pack", title_style),
        Paragraph("Professional policy export for audit-ready review and approval.", subtitle_style),
        Spacer(1, 0.12 * inch),
    ]

    cover_metadata = [
        ["Client", clean_inline_text(client_name)],
        ["Generated By", "DB Agent"],
        ["Export Timestamp", clean_inline_text(generated_at)],
        ["Export Scope", clean_inline_text(scope_text)],
        ["Policy Count", str(len(policies))],
    ]
    cover_table = Table(cover_metadata, colWidths=[1.5 * inch, 4.7 * inch], hAlign="LEFT")
    cover_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E2E8F0")),
                ("BACKGROUND", (1, 0), (1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEADING", (0, 0), (-1, -1), 12),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([cover_table, Spacer(1, 0.2 * inch)])

    if not policies:
        story.append(Paragraph("No policies were available for export.", body_style))
    else:
        for index, policy in enumerate(policies):
            metadata_rows = [
                ["Policy Name", clean_inline_text(policy.get("name") or f"Policy {index + 1}")],
                ["Policy ID", clean_inline_text(policy.get("policy_id"))],
                ["Version", clean_inline_text(policy.get("policy_version") or "v1.0")],
                ["Owner", clean_inline_text(policy.get("policy_owner")) or "Pending assignment"],
                ["Approver", clean_inline_text(policy.get("sign_off_by")) or "Pending assignment"],
                ["Published", clean_inline_text(policy.get("published")) or "No"],
                ["Published At", clean_inline_text(policy.get("published_at")) or "Not published"],
                ["Signed Off", clean_inline_text(policy.get("sign_off_complete")) or "No"],
                ["Signed Off At", clean_inline_text(policy.get("sign_off_completed_at")) or "Not signed off"],
            ]

            story.append(Paragraph(escape(clean_inline_text(policy.get("name") or f"Policy {index + 1}")), policy_title_style))
            policy_table = Table(metadata_rows, colWidths=[1.4 * inch, 4.8 * inch], hAlign="LEFT")
            policy_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F8FAFC")),
                        ("BACKGROUND", (1, 0), (1, -1), colors.white),
                        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 7),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.extend([policy_table, Spacer(1, 0.12 * inch)])

            if policy.get("executive_summary"):
                story.append(Paragraph("Executive Summary", section_heading_style))
                story.append(Paragraph(escape(clean_inline_text(policy["executive_summary"])), body_style))
            if policy.get("table_of_contents"):
                story.append(Paragraph("Table of Contents", section_heading_style))
                for toc_line in sanitize_text(policy["table_of_contents"]).replace("\r", "").split("\n"):
                    toc_line = toc_line.strip()
                    if toc_line:
                        story.append(Paragraph(escape(toc_line), muted_style))
            if policy.get("body"):
                story.append(Paragraph("Policy Document", section_heading_style))
                add_policy_text_block(story, policy["body"])
            if policy.get("approval_history_text"):
                story.append(Paragraph("Approval History", section_heading_style))
                for history_line in sanitize_text(policy["approval_history_text"]).replace("\r", "").split("\n"):
                    history_line = history_line.strip()
                    if history_line:
                        story.append(Paragraph(escape(history_line), muted_style))
            if index < len(policies) - 1:
                story.append(PageBreak())

    document = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        title="DB Agent Policy Pack",
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.65 * inch,
    )
    document.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return str(path.resolve())


def build_policy_bundle(export_root: Path, client_name: str, policies: list[dict], generated_at: str, scope_text: str) -> tuple[str, str]:
    bundle_root = ensure_directory(export_root / "policy-pdfs")
    generated_files: list[Path] = []

    for index, policy in enumerate(policies, start=1):
        file_name = f"{index:02d}-{slugify(sanitize_text(policy.get('name') or policy.get('policy_id') or 'policy'))}.pdf"
        file_path = bundle_root / file_name
        build_policy_pdf(file_path, client_name, [policy], generated_at, scope_text)
        generated_files.append(file_path)

    zip_path = export_root / "policy-pdfs.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in generated_files:
            archive.write(file_path, arcname=f"policy-pdfs/{file_path.name}")

    return str(zip_path.resolve()), str(bundle_root.resolve())


def write_text_report(path: Path, heading: str, records: list[dict], generated_at: str, scope_text: str) -> str:
    lines = [
        heading,
        "Generated by DB Agent",
        f"Export timestamp: {generated_at}",
        f"Export scope: {scope_text}",
        "",
    ]
    if not records:
        lines.append("No records available.")
    else:
        for record in records:
            lines.append(f"- {sanitize_text(record)}")
    path.write_text("\n".join(lines), encoding="utf-8")
    return str(path.resolve())


def rows_from_collection(items: Iterable[dict], headers: list[str]) -> list[dict]:
    rows: list[dict] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        rows.append({header: item.get(header, "") for header in headers})
    return rows


def main() -> None:
    args = parse_args()
    client_root = Path(args.client_root)
    export_root = ensure_directory(Path(args.export_root))
    generated_at = now_iso()

    onboarding = read_json(client_section_path(client_root, "onboarding"), {})
    policy_generation = read_json(client_section_path(client_root, "policy_generation"), {})
    risk_assessment = read_json(client_section_path(client_root, "risk_assessment"), {})
    vendor_risk = read_json(client_section_path(client_root, "vendor_risk"), {})
    control_mapping = read_json(client_section_path(client_root, "control_mapping"), {})
    audit_qa = read_json(client_section_path(client_root, "audit_qa"), {})

    client_name = sanitize_text(onboarding.get("legal_entity") or args.client_id)
    scope_text = sanitize_text(onboarding.get("scope") or onboarding.get("framework_selection") or args.client_id)

    policies = list(policy_generation.get("policies") or [])
    risks = list(risk_assessment.get("risks") or [])
    vendors = list(vendor_risk.get("vendors") or [])
    controls = list(control_mapping.get("controls") or [])
    audit_findings = list(audit_qa.get("findings") or [])

    risk_headers = [
        "risk_id",
        "title",
        "asset",
        "process_area",
        "threat",
        "vulnerability",
        "impact_description",
        "likelihood",
        "impact",
        "inherent_score",
        "residual_likelihood",
        "residual_impact",
        "residual_score",
        "treatment_plan",
        "owner",
        "linked_policies",
        "linked_controls",
    ]
    vendor_headers = [
        "vendor_id",
        "vendor_name",
        "vendor_description",
        "purpose",
        "service_category",
        "known_services",
        "website",
        "service_type",
        "business_owner",
        "security_owner",
        "data_access_level",
        "data_types",
        "criticality",
        "vendor_likelihood",
        "vendor_impact",
        "inherent_score",
        "inherent_risk",
        "residual_likelihood",
        "residual_impact",
        "residual_score",
        "residual_risk",
        "assessment_status",
        "treatment_plan",
        "linked_risks",
        "linked_controls",
        "final_decision",
    ]
    control_headers = [
        "control_id",
        "description",
        "owner",
        "frequency",
        "evidence",
        "linked_policies",
        "linked_risks",
        "linked_vendors",
        "framework_mapping",
    ]

    outputs = []

    policy_pdf = build_policy_pdf(export_root / "policy-pack.pdf", client_name, policies, generated_at, scope_text)
    policy_zip, _policy_folder = build_policy_bundle(export_root, client_name, policies, generated_at, scope_text)
    outputs.append(
        {
            "output_id": "OUT-001",
            "output_type": "Policy pack",
            "files": [policy_pdf, policy_zip],
        }
    )

    risk_rows = rows_from_collection(risks, risk_headers)
    risk_csv = write_csv(export_root / "risk-register.csv", risk_rows, risk_headers)
    risk_xlsx = write_xlsx(export_root / "risk-register.xlsx", risk_rows, risk_headers, "Risk Register")
    outputs.append(
        {
            "output_id": "OUT-002",
            "output_type": "Risk register",
            "files": [risk_csv, risk_xlsx],
        }
    )

    vendor_rows = rows_from_collection(vendors, vendor_headers)
    vendor_csv = write_csv(export_root / "vendor-register.csv", vendor_rows, vendor_headers)
    vendor_xlsx = write_xlsx(export_root / "vendor-register.xlsx", vendor_rows, vendor_headers, "Vendor Register")
    outputs.append(
        {
            "output_id": "OUT-003",
            "output_type": "Vendor register",
            "files": [vendor_csv, vendor_xlsx],
        }
    )

    control_rows = rows_from_collection(controls, control_headers)
    control_csv = write_csv(export_root / "control-library.csv", control_rows, control_headers)
    control_xlsx = write_xlsx(export_root / "control-library.xlsx", control_rows, control_headers, "Control Library")
    outputs.append(
        {
            "output_id": "OUT-004",
            "output_type": "Control library",
            "files": [control_csv, control_xlsx],
        }
    )

    audit_txt = write_text_report(export_root / "audit-readiness.txt", "Audit Readiness Report", audit_findings, generated_at, scope_text)
    outputs.append(
        {
            "output_id": "OUT-005",
            "output_type": "Audit readiness report",
            "files": [audit_txt],
        }
    )

    manifest = {
        "status": "success",
        "generated_at": generated_at,
        "export_scope": scope_text,
        "client_id": args.client_id,
        "client_name": client_name,
        "generated_by": "DB Agent",
        "outputs": outputs,
    }
    manifest_path = export_root / "export-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    for output in outputs:
        if output["output_id"] == "OUT-005":
            output["files"].append(str(manifest_path.resolve()))

    print(json.dumps(manifest, ensure_ascii=False))


if __name__ == "__main__":
    main()
